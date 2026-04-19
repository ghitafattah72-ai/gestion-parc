"""
Utilities for Excel/CSV export operations
"""
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import send_file


def _autosize_worksheet_columns(ws, max_width=60):
    """Set a readable width for each used column based on content length."""
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)


def export_to_csv(headers, rows, filename):
    """
    Export data to CSV format
    
    Args:
        headers: List of header strings
        rows: List of data rows (each row should be a list/tuple)
        filename: Name of the file to download
    """
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )


def export_to_excel(headers, rows, filename, sheet_name='Data'):
    """
    Export data to Excel format
    
    Args:
        headers: List of header strings
        rows: List of data rows (each row should be a list/tuple)
        filename: Name of the file to download
        sheet_name: Name of the Excel sheet
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add headers
    ws.append(headers)
    
    # Add data rows
    for row in rows:
        ws.append(row)

    _autosize_worksheet_columns(ws)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_to_excel_sheets(sheets, filename):
    """
    Export multiple sheets to a single Excel file.

    Args:
        sheets: List of dicts with keys: sheet_name, headers, rows
        filename: Name of the file to download
    """
    wb = Workbook()

    # Remove default sheet so we can control the order/names explicitly.
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.get('sheet_name', 'Data'))
        headers = sheet.get('headers', [])
        rows = sheet.get('rows', [])

        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)

        _autosize_worksheet_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def get_export_filename(base_name, format_type='csv'):
    """
    Generate a filename with timestamp
    
    Args:
        base_name: Base name of the file (e.g., 'stock')
        format_type: File format ('csv' or 'xlsx')
    
    Returns:
        Filename with timestamp
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    extension = 'xlsx' if format_type == 'xlsx' else 'csv'
    return f'{base_name}_{timestamp}.{extension}'
